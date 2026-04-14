"""
historical.py — Ingest historical capacity data from IAEA PRIS RDS-2.

Table 5  — Country capacity at 5-year snapshots: 1995,2000,2005,2010,2015,2020,2023,2024
           Layout (0-based): 0=None, 1=Country, then pairs per year:
           2=No(1995), 3=MW(1995), 4=No(2000), 5=MW(2000), ..., 16=No(2024), 17=MW(2024)

Table 7  — Global annual data 1954–2024.
           Layout (0-based): 0=None, 1=Year(str), 2=CS_Units, 3=CS_MW,
           4=Grid_Units, 5=Grid_MW, 6=Op_Units, 7=Op_MW
"""
import sqlite3
from datetime import date
from pathlib import Path
import sys
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from config import RAW_PRIS_DIR, REGION_MAP, HISTORICAL_START_YEAR

import openpyxl


def pris_path(n: int) -> Path:
    return RAW_PRIS_DIR / f"2024_Table{n}.xlsx"


def get_region(country: str) -> str:
    return REGION_MAP.get(country.strip().upper(), "Emerging & Rest")


def load_table5_country_capacity() -> list[dict]:
    """
    Parse Table 5 for country-level historical capacity.
    Years (column pairs starting at col 2): 1995,2000,2005,2010,2015,2020,2023,2024
    Each year has 2 columns: No. (col 2i) and MW(e) (col 2i+1), starting at i=1.

    Column layout:
      Col 0: None
      Col 1: Country
      Col 2: No.(1995), Col 3: MW(1995)
      Col 4: No.(2000), Col 5: MW(2000)
      Col 6: No.(2005), Col 7: MW(2005)
      Col 8: No.(2010), Col 9: MW(2010)
      Col 10: No.(2015), Col 11: MW(2015)
      Col 12: No.(2020), Col 13: MW(2020)
      Col 14: No.(2023), Col 15: MW(2023)
      Col 16: No.(2024), Col 17: MW(2024)
    """
    wb = openpyxl.load_workbook(pris_path(5), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Confirm year positions from the year header row (row index 5 = row 6)
    year_row = rows[5]  # [None, '', 1995, None, 2000, None, 2005, ...]
    # Years appear at positions 2,4,6,8,10,12,14,16 (every other starting at 2)
    year_positions: list[tuple[int, int]] = []  # (year, no_col) — mw_col = no_col+1
    for i, val in enumerate(year_row):
        if isinstance(val, (int, float)) and 1990 <= val <= 2030:
            year_positions.append((int(val), i))

    records = []
    for row in rows[7:]:  # data rows start at row 8 (0-based 7)
        if all(v is None for v in row):
            continue
        country_raw = row[1] if len(row) > 1 else None
        if not country_raw or not str(country_raw).strip():
            continue
        country = str(country_raw).strip().upper()
        if country in ("", "TOTAL", "WORLD TOTAL"):
            continue
        region = get_region(country)

        for yr, no_col in year_positions:
            if yr < HISTORICAL_START_YEAR:
                continue
            mw_col = no_col + 1
            num_r = row[no_col]  if no_col  < len(row) else None
            mw    = row[mw_col] if mw_col < len(row) else None
            if mw is not None:
                try:
                    records.append({
                        "year":         yr,
                        "country":      country,
                        "region":       region,
                        "capacity_gw":  round(float(mw) / 1000.0, 4),
                        "num_reactors": int(num_r) if num_r and str(num_r).strip() else 0,
                    })
                except (ValueError, TypeError):
                    pass
    return records


def load_table7_global_annual() -> list[dict]:
    """
    Parse Table 7 for global annual data.
    Layout (0-based): 0=None, 1=Year(string!), 2=CS_Units, 3=CS_MW,
    4=Grid_Units, 5=Grid_MW, 6=Op_Units, 7=Op_MW
    Year is stored as a string e.g. '1954', '2024'.
    """
    wb = openpyxl.load_workbook(pris_path(7), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    records = []
    for row in rows[6:]:   # data rows at row 7 (0-based 6)
        if all(v is None for v in row):
            continue
        year_raw = row[1] if len(row) > 1 else None
        if year_raw is None:
            continue
        # Year is stored as string '1954' etc.
        year_str = str(year_raw).strip()
        if not year_str.isdigit():
            continue
        yr = int(year_str)
        if yr < HISTORICAL_START_YEAR:
            continue
        op_units = row[6] if len(row) > 6 else None
        op_mw    = row[7] if len(row) > 7 else None
        if op_mw is not None:
            try:
                records.append({
                    "year":         yr,
                    "capacity_gw":  round(float(op_mw) / 1000.0, 4),
                    "num_reactors": int(op_units) if op_units else 0,
                })
            except (ValueError, TypeError):
                pass
    return records


def ingest_historical(conn: sqlite3.Connection) -> None:
    """Load historical capacity data."""
    print("  Loading Table 5 (country historical capacity)...")
    country_records = load_table5_country_capacity()
    print(f"    → {len(country_records)} country-year records")

    print("  Loading Table 7 (global annual capacity 2005–2024)...")
    global_records = load_table7_global_annual()
    print(f"    → {len(global_records)} global annual records")

    cur = conn.cursor()

    # Aggregate country records to region level
    region_year: dict[tuple, dict] = defaultdict(lambda: {"capacity_gw": 0.0, "num_reactors": 0})
    for rec in country_records:
        key = (rec["year"], rec["region"])
        region_year[key]["capacity_gw"]  += rec["capacity_gw"]
        region_year[key]["num_reactors"] += rec["num_reactors"]

    # Insert regional records (5-yr snapshots)
    inserted = 0
    for (yr, region), vals in region_year.items():
        cur.execute("""
            INSERT OR REPLACE INTO historical_capacity (year, region, capacity_gw, num_reactors, source)
            VALUES (?, ?, ?, ?, ?)
        """, (yr, region, round(vals["capacity_gw"], 3), vals["num_reactors"], "IAEA_RDS2_T5"))
        inserted += 1

    # Insert Global totals from Table 7 (annual granularity 2005–2024)
    for rec in global_records:
        cur.execute("""
            INSERT OR REPLACE INTO historical_capacity (year, region, capacity_gw, num_reactors, source)
            VALUES (?, ?, ?, ?, ?)
        """, (rec["year"], "Global", round(rec["capacity_gw"], 3), rec["num_reactors"], "IAEA_RDS2_T7"))
        inserted += 1

    conn.commit()
    print(f"  ✓ Inserted {inserted} historical capacity records")

    cur.execute("""
        INSERT OR REPLACE INTO data_sources_log (source_name, data_as_of_date, ingestion_date, version_note)
        VALUES (?, ?, ?, ?)
    """, ("IAEA_RDS2_Historical", "2024-12-31", str(date.today()),
          "IAEA RDS-2 2024 — Table 5 (country 5yr snapshots) and Table 7 (global annual)"))
    conn.commit()
