"""
cache.py — Streamlit-cached data loaders.

All @st.cache_data wrappers live here so they can be shared between
app.py and any tab renderer without re-importing streamlit in multiple
places.  The 5-minute TTL matches the DB update cadence.
"""
from __future__ import annotations

import io
import json
import sys
from pathlib import Path

import pandas as pd
import streamlit as st

ROOT = Path(__file__).parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from config import REGIONS
from model.api import (
    get_projection,
    get_historical,
    get_benchmarks,
    get_reactor_list,
    get_data_vintage_string,
    get_country_capacity,
    get_full_reactor_download,
    get_tech_projection,
)
from model.state import ProjectionState, build_projection_state


@st.cache_data(ttl=300)
def load_all_projections() -> dict[str, dict[str, pd.DataFrame]]:
    scenario_ids = ["decline", "conservative", "base", "optimistic"]
    result: dict[str, dict[str, pd.DataFrame]] = {}
    for sc_id in scenario_ids:
        result[sc_id] = {}
        for region in REGIONS + ["Global"]:
            result[sc_id][region] = get_projection(scenario_id=sc_id, region=region)
    return result


@st.cache_data(ttl=300)
def load_historical_all() -> dict[str, pd.DataFrame]:
    result: dict[str, pd.DataFrame] = {}
    for region in REGIONS + ["Global"]:
        result[region] = get_historical(region=region)
    return result


@st.cache_data(ttl=300)
def load_benchmarks() -> pd.DataFrame:
    return get_benchmarks(region="Global")


@st.cache_data(ttl=300)
def load_reactors() -> pd.DataFrame:
    return get_reactor_list()


@st.cache_data(ttl=600)
def load_vintage() -> str:
    return get_data_vintage_string()


@st.cache_data(ttl=600)
def load_full_reactor_download() -> pd.DataFrame:
    return get_full_reactor_download()


@st.cache_data(ttl=300, show_spinner=False)
def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Data") -> bytes:
    """Convert a DataFrame to styled Excel bytes for st.download_button.
    Cached so repeated re-runs (slider moves, tab switches) don't regenerate the file."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        from openpyxl.styles import Font, PatternFill, Alignment
        hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        ws.freeze_panes = "A2"
    return buf.getvalue()


@st.cache_data(ttl=300, show_spinner=False)
def load_projection_state(
    scenario_id: str,
    what_if_overrides_json: str = "",
) -> ProjectionState:
    """
    Cached wrapper for build_projection_state.

    Uses a JSON string as the cache key (ProjectionState itself is not
    hashable by st.cache_data).  Called once per (scenario_id, overrides)
    pair per session; the result is shared across all tab consumers.
    """
    wi_overrides = json.loads(what_if_overrides_json) if what_if_overrides_json else None
    return build_projection_state(scenario_id, wi_overrides)


@st.cache_data(ttl=300, show_spinner=False)
def load_tech_projection(
    scenario_id: str,
    regions_key: str,           # stringified tuple of regions — hashable cache key
    smr_post2040_share: float,
    smr_accel_start_year: int = 0,
    smr_accel_gw_per_year: float = 0.0,
    what_if_overrides_json: str = "",
) -> pd.DataFrame:
    """
    Cached wrapper for get_tech_projection.

    When what_if_overrides_json is set, reuses the already-cached
    ProjectionState so DB reads and override application happen at most
    once per (scenario, overrides) pair.
    """
    regions = list(eval(regions_key)) if regions_key != "all" else None
    state = (
        load_projection_state(scenario_id, what_if_overrides_json)
        if what_if_overrides_json else None
    )
    return get_tech_projection(
        scenario_id=scenario_id,
        regions=regions,
        smr_post2040_share=smr_post2040_share,
        smr_accel_start_year=smr_accel_start_year,
        smr_accel_gw_per_year=smr_accel_gw_per_year,
        what_if_overrides=None,   # never pass raw overrides — state handles them
        state=state,
    )


@st.cache_data(ttl=300, show_spinner=False)
def load_country_capacity(year: int, scenario_id: str) -> pd.DataFrame:
    """Cached wrapper for get_country_capacity."""
    return get_country_capacity(year=year, scenario_id=scenario_id)
