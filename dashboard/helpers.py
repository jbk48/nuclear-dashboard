"""
helpers.py — Pure helper functions for the dashboard.

No @st.cache_data decorators here; some functions read st.session_state.
Imported by both app.py and tab_renderers.py.
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import pandas as pd
import streamlit as st

ROOT = Path(__file__).parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from dashboard.overrides import WhatIfOverrides
from dashboard.levers import (
    ScenarioState, PRESET_DEFAULTS,
    LARGE_PIPELINE_PRESETS, SMR_PIPELINE_PRESETS,
)

# ── smr_post2040_share is stored as integer percent in levers.py ──────────────
# Convert once at import time so levers_match_preset() uses float fractions.
_PRESET_DEFAULTS_FLOAT: dict = {
    k: {**v, "smr_post2040_share": v["smr_post2040_share"] / 100.0}
    for k, v in PRESET_DEFAULTS.items()
}


# ── Override dict builder ─────────────────────────────────────────────────────

def build_wi_overrides_dict() -> dict:
    """Rebuild the what_if_overrides raw dict from session state for model API calls."""
    overrides = WhatIfOverrides.from_session_state(
        st.session_state.get("wi_overrides", []),
        st.session_state.get("wi_synthetic", []),
    )
    return overrides.to_raw_dict()


# ── Geography aggregation ─────────────────────────────────────────────────────

def sum_projections(proj_dict: dict, regions: list) -> pd.DataFrame:
    """Sum per-region projection DataFrames across the given regions.
    Uses year-based groupby so misaligned or differently-sized DataFrames
    cannot cause silent positional errors."""
    sum_cols = [
        "capacity_operating_gw", "retirements_this_year_gw", "additions_this_year_gw",
        "capacity_retired_ytd_gw", "capacity_added_ytd_gw",
    ]
    sample = next(
        (proj_dict[r] for r in regions if r in proj_dict and not proj_dict[r].empty), None
    )
    if sample is None:
        return pd.DataFrame()
    available_sum_cols = [c for c in sum_cols if c in sample.columns]
    dfs = [
        proj_dict[r][["year", "is_bottom_up"] + available_sum_cols]
        for r in regions if r in proj_dict and not proj_dict[r].empty
    ]
    if not dfs:
        return pd.DataFrame()
    combined = pd.concat(dfs)
    ib      = combined.groupby("year")["is_bottom_up"].max().reset_index()
    summed  = combined.groupby("year")[available_sum_cols].sum().reset_index()
    return summed.merge(ib, on="year")


def sum_historical(hist_dict: dict, regions: list) -> pd.DataFrame:
    """Sum per-region historical DataFrames across the given regions."""
    dfs = [
        hist_dict[r][["year", "capacity_gw"]]
        for r in regions if r in hist_dict and not hist_dict[r].empty
    ]
    if not dfs:
        return pd.DataFrame()
    return pd.concat(dfs).groupby("year")["capacity_gw"].sum().reset_index()


def gw_from_df(df: pd.DataFrame, default: float = 0.0) -> float:
    """Extract capacity_operating_gw scalar from a single-row filtered DataFrame."""
    return float(df["capacity_operating_gw"].iloc[0]) if not df.empty else default


# ── Scenario comparison ───────────────────────────────────────────────────────

def levers_match_preset(state: ScenarioState, preset_id: str) -> bool:
    pd_ = _PRESET_DEFAULTS_FLOAT.get(preset_id, {})
    return (
        state.extension_policy_global == pd_.get("extension_policy_global")
        and state.large_pipeline_preset == pd_.get("large_pipeline_preset")
        and state.smr_pipeline_preset == pd_.get("smr_pipeline_preset")
        and abs(state.construction_delay_adder - pd_.get("construction_delay_adder", 0)) < 0.01
        and abs(state.smr_accel_gw_per_year - pd_.get("smr_accel_gw_per_year", 0)) < 0.01
        and abs(state.smr_post2040_share - pd_.get("smr_post2040_share", 0.20)) < 0.01
        and abs(state.post2040_global_growth_gw - pd_.get("post2040_global_growth_gw", 0)) < 0.1
        and abs(state.china_post2040_gw - pd_.get("china_post2040_gw", 0)) < 0.1
    )


def build_scenario_bullets(
    state: ScenarioState,
    sc_id: str,
    wi_overrides: list,
    wi_synthetic: list,
) -> list[str]:
    """
    Generate brief directional impact bullets for the Scenario Lab diff panel.
    Compares the current ScenarioState against the starting preset (sc_id),
    then summarises any reactor overrides and synthetic builds.
    ⬆️ = growth driver / accelerator   ⬇️ = downward pressure / delayer
    """
    bullets: list[str] = []
    pd_ = PRESET_DEFAULTS.get(sc_id, PRESET_DEFAULTS["base"])

    # ── Life extension policy ─────────────────────────────────────────────────
    ext_def  = pd_.get("extension_policy_global", "CurrentPolicy")
    ext_curr = state.extension_policy_global
    if ext_curr != ext_def:
        _ext_map = {
            "ExtendedOperations":    ("⬆️", "Extended Operations — all reactors run to regulatory max life, expanding the operating fleet"),
            "AcceleratedRetirement": ("⬇️", "Accelerated Retirement — no new licence extensions, fleet declines faster"),
            "CurrentPolicy":         ("➡️", "Current Policy — reverted to country-specific licensing rules"),
        }
        arrow, desc = _ext_map.get(ext_curr, ("~", ext_curr))
        bullets.append(f"{arrow} **Reactor life extension**: {desc}")

    # ── Large reactor pipeline ────────────────────────────────────────────────
    lpp     = LARGE_PIPELINE_PRESETS[pd_.get("large_pipeline_preset", "medium")]
    luc_def = int(lpp["uc_rate"]       * 100)
    lpl_def = int(lpp["planned_rate"]  * 100)
    lpr_def = int(lpp["proposed_rate"] * 100)
    luc_cur = int(state.pipeline_uc_rate       * 100)
    lpl_cur = int(state.pipeline_planned_rate  * 100)
    lpr_cur = int(state.pipeline_proposed_rate * 100)
    if luc_cur != luc_def or lpl_cur != lpl_def or lpr_cur != lpr_def:
        score = (luc_cur - luc_def) + (lpl_cur - lpl_def) + (lpr_cur - lpr_def)
        arrow = "⬆️" if score > 0 else "⬇️"
        bullets.append(
            f"{arrow} **Large reactor pipeline**: UC {luc_cur}% / Planned {lpl_cur}% / Proposed {lpr_cur}%"
            f" (default {luc_def}% / {lpl_def}% / {lpr_def}%)"
        )

    # ── SMR pipeline ──────────────────────────────────────────────────────────
    spp     = SMR_PIPELINE_PRESETS[pd_.get("smr_pipeline_preset", "medium")]
    suc_def = int(spp["uc_rate"]       * 100)
    spl_def = int(spp["planned_rate"]  * 100)
    spr_def = int(spp["proposed_rate"] * 100)
    suc_cur = int(state.smr_uc_rate       * 100)
    spl_cur = int(state.smr_planned_rate  * 100)
    spr_cur = int(state.smr_proposed_rate * 100)
    if suc_cur != suc_def or spl_cur != spl_def or spr_cur != spr_def:
        score = (suc_cur - suc_def) + (spl_cur - spl_def) + (spr_cur - spr_def)
        arrow = "⬆️" if score > 0 else "⬇️"
        bullets.append(
            f"{arrow} **SMR pipeline**: UC {suc_cur}% / Planned {spl_cur}% / Proposed {spr_cur}%"
            f" (default {suc_def}% / {spl_def}% / {spr_def}%)"
        )

    # ── Construction delay ────────────────────────────────────────────────────
    delay_def = pd_.get("construction_delay_adder", 0)
    delay_cur = state.construction_delay_adder
    if abs(delay_cur - delay_def) >= 0.5:
        if delay_cur > delay_def:
            bullets.append(
                f"⬇️ **Construction delay**: +{delay_cur:.0f} yr"
                f" (default +{delay_def:.0f} yr) — all pipeline additions pushed back"
            )
        else:
            bullets.append(
                f"⬆️ **Construction delay**: +{delay_cur:.0f} yr"
                f" (default +{delay_def:.0f} yr) — pipeline additions pulled forward"
            )

    # ── SMR pre-2040 acceleration ─────────────────────────────────────────────
    accel_def = float(pd_.get("smr_accel_gw_per_year", 0))
    accel_cur = state.smr_accel_gw_per_year
    if accel_cur > max(accel_def, 0.05):
        bullets.append(
            f"⬆️ **SMR pre-2040 acceleration**: +{accel_cur:.1f} GW/yr"
            f" from {state.smr_accel_start_year} — extra SMR capacity above announced pipeline"
        )
    elif accel_def > 0.05 and accel_cur < accel_def - 0.05:
        bullets.append(
            f"⬇️ **SMR acceleration reduced**: {accel_cur:.1f} GW/yr (default {accel_def:.1f})"
        )

    # ── Post-2040 global new build ────────────────────────────────────────────
    p40_def = float(pd_.get("post2040_global_growth_gw", 28.1))
    p40_cur = state.post2040_global_growth_gw
    if abs(p40_cur - p40_def) > 0.5:
        if p40_cur > p40_def:
            bullets.append(
                f"⬆️ **Post-2040 new build**: {p40_cur:.1f} GW/yr"
                f" (default {p40_def:.1f}) — higher long-run capacity growth"
            )
        else:
            bullets.append(
                f"⬇️ **Post-2040 new build**: {p40_cur:.1f} GW/yr"
                f" (default {p40_def:.1f}) — lower long-run capacity growth"
            )

    # ── China post-2040 share ─────────────────────────────────────────────────
    china_def = float(pd_.get("china_post2040_gw", 8.0))
    china_cur = state.china_post2040_gw
    if abs(china_cur - china_def) > 0.5:
        arrow = "⬆️" if china_cur > china_def else "⬇️"
        bullets.append(
            f"{arrow} **China post-2040 share**: {china_cur:.1f} GW/yr"
            f" (default {china_def:.1f}) — shifts regional mix"
        )

    # ── Reactor overrides ─────────────────────────────────────────────────────
    if wi_overrides:
        cancelled   = [o for o in wi_overrides
                       if o["field"] == "pipeline_probability" and float(o["value"]) < 0.1]
        confirmed   = [o for o in wi_overrides
                       if o["field"] == "pipeline_probability" and float(o["value"]) >= 0.9]
        ret_changes = [o for o in wi_overrides if o["field"] == "retirement_year"]
        restarts    = [o for o in wi_overrides if o["field"] == "restart_date"]
        cap_changes = [o for o in wi_overrides if o["field"] == "capacity_mw"]
        other_ov    = [o for o in wi_overrides
                       if o["field"] not in
                       {"pipeline_probability", "retirement_year", "restart_date", "capacity_mw"}]

        if cancelled:
            bullets.append(
                f"⬇️ **{len(cancelled)} pipeline reactor(s) cancelled**"
                " — probability → 0, removing those additions from the projection"
            )
        if confirmed:
            bullets.append(
                f"⬆️ **{len(confirmed)} pipeline reactor(s) confirmed**"
                " — probability → 1, locking in those additions"
            )
        if ret_changes:
            yrs = [float(o["value"]) for o in ret_changes]
            avg_yr = sum(yrs) / len(yrs)
            if avg_yr >= 2045:
                bullets.append(
                    f"⬆️ **{len(ret_changes)} reactor life extension(s)**"
                    f" — retirement year avg ~{avg_yr:.0f}, fleet lives longer"
                )
            elif avg_yr <= 2034:
                bullets.append(
                    f"⬇️ **{len(ret_changes)} early retirement(s)**"
                    f" — closure year avg ~{avg_yr:.0f}, capacity removed sooner"
                )
            else:
                bullets.append(
                    f"~ **{len(ret_changes)} retirement year adjustment(s)**"
                    f" — avg target ~{avg_yr:.0f}"
                )
        if restarts:
            bullets.append(
                f"⬆️ **{len(restarts)} reactor restart(s)**"
                " — previously shutdown units returned to service"
            )
        if cap_changes:
            bullets.append(
                f"~ **{len(cap_changes)} reactor capacity adjustment(s)**"
                " — net rating changed on specific units"
            )
        if other_ov:
            bullets.append(f"~ **{len(other_ov)} other reactor override(s)**")

    # ── Synthetic new builds ──────────────────────────────────────────────────
    for sb in wi_synthetic:
        gw_yr = sb["capacity_mw"] * sb["per_year"] / 1000
        bullets.append(
            f"⬆️ **Synthetic build**: {sb['per_year']} × {int(sb['capacity_mw'])} MW/yr"
            f" in **{sb['region']}** from {sb['start_year']}"
            f" for {sb['n_years']} yr (+{gw_yr:.1f} GW/yr)"
        )

    return bullets


# ── Scenario export ───────────────────────────────────────────────────────────

def build_scenario_export(
    state: ScenarioState,
    active_key: str,
    proj_global: pd.DataFrame,
    proj_by_region: dict,
) -> bytes:
    """
    Build a multi-sheet Excel export for the current scenario:
      Sheet 1 — Scenario Settings  (all lever values, fully reproducible)
      Sheet 2 — Global Projection  (year-by-year global totals)
      Sheet 3 — Regional Breakdown (year-by-year per region)
      Sheet 4 — About              (model description + disclaimer)
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import date as _date

    buf = io.BytesIO()

    def _style_sheet(ws, header_color: str = "1F4E79") -> None:
        hdr_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
        ws.freeze_panes = "A2"

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── Sheet 1: Scenario Settings ────────────────────────────────────────
        ext_labels = {
            "None":           "None — retire at baseline date",
            "HistoricalRate": "Historical Rate (~50% of fleet extended)",
            "Moderate":       "Moderate (~75% of fleet extended)",
            "MaximumAllowed": "Maximum Allowed (100%, max regulatory life)",
        }
        pipeline_desc = {
            "high":   "High — UC + Planned + Proposed all complete",
            "medium": "Medium — UC + Planned complete; Proposed excluded",
            "low":    "Low — Under Construction only",
        }
        settings_rows = [
            ("Scenario",                  active_key.title()),
            ("Export date",               str(_date.today())),
            ("",                          ""),
            ("── Life Extensions",        ""),
            ("Extension policy",          ext_labels.get(state.extension_policy_global,
                                                          state.extension_policy_global)),
            ("",                          ""),
            ("── Pipeline Realization",   ""),
            ("Large reactor preset",      pipeline_desc.get(state.large_pipeline_preset,
                                                             state.large_pipeline_preset)),
            ("  Under Construction rate", f"{state.pipeline_uc_rate*100:.0f}%"),
            ("  Planned rate",            f"{state.pipeline_planned_rate*100:.0f}%"),
            ("  Proposed rate",           f"{state.pipeline_proposed_rate*100:.0f}%"),
            ("SMR pipeline preset",       pipeline_desc.get(state.smr_pipeline_preset,
                                                             state.smr_pipeline_preset)),
            ("  SMR Under Construction",  f"{state.smr_uc_rate*100:.0f}%"),
            ("  SMR Planned",             f"{state.smr_planned_rate*100:.0f}%"),
            ("  SMR Proposed",            f"{state.smr_proposed_rate*100:.0f}%"),
            ("Construction delay adder",  f"+{state.construction_delay_adder:.0f} years"),
            ("",                          ""),
            ("── SMR Deployment",         ""),
            ("Pre-2040 SMR acceleration",
             f"{state.smr_accel_gw_per_year:.1f} GW/yr"
             + (f" from {state.smr_accel_start_year}" if state.smr_accel_gw_per_year > 0 else " (none)")),
            ("Post-2040 SMR share",       f"{state.smr_post2040_share*100:.0f}%"),
            ("",                          ""),
            ("── Post-2040 New Build",    ""),
            ("Global new build rate",     f"{state.post2040_global_growth_gw:.1f} GW/yr"),
            ("China share",               f"{state.china_post2040_gw:.1f} GW/yr"),
            ("Rest-of-world share",
             f"{max(0.0, state.post2040_global_growth_gw - state.china_post2040_gw):.1f} GW/yr"),
        ]
        df_settings = pd.DataFrame(settings_rows, columns=["Setting", "Value"])
        df_settings.to_excel(writer, index=False, sheet_name="Scenario Settings")
        ws_s = writer.sheets["Scenario Settings"]
        _style_sheet(ws_s, header_color="1F4E79")
        for row in ws_s.iter_rows(min_row=2):
            if str(row[0].value or "").startswith("──"):
                for cell in row:
                    cell.font = Font(bold=True, color="1F4E79")

        # ── Sheet 2: Global Projection ────────────────────────────────────────
        _base_cols = ["year", "capacity_operating_gw",
                      "retirements_this_year_gw", "additions_this_year_gw", "is_bottom_up"]
        _ytd_cols  = ["capacity_retired_ytd_gw", "capacity_added_ytd_gw"]
        _avail_ytd = [c for c in _ytd_cols if c in proj_global.columns]
        df_global  = proj_global[_base_cols[:4] + _avail_ytd + _base_cols[4:]].copy()
        _col_names = ["Year", "Capacity (GW)", "Retirements (GW/yr)", "Additions (GW/yr)"]
        if "capacity_retired_ytd_gw" in _avail_ytd:
            _col_names.append("Cumul. Retired (GW)")
        if "capacity_added_ytd_gw" in _avail_ytd:
            _col_names.append("Cumul. Added (GW)")
        _col_names.append("Bottom-Up Phase")
        df_global.columns = _col_names
        df_global["Capacity (GW)"]       = df_global["Capacity (GW)"].round(2)
        df_global["Retirements (GW/yr)"] = df_global["Retirements (GW/yr)"].round(3)
        df_global["Additions (GW/yr)"]   = df_global["Additions (GW/yr)"].round(3)
        if "Cumul. Retired (GW)" in df_global.columns:
            df_global["Cumul. Retired (GW)"] = df_global["Cumul. Retired (GW)"].round(2)
        if "Cumul. Added (GW)" in df_global.columns:
            df_global["Cumul. Added (GW)"]   = df_global["Cumul. Added (GW)"].round(2)
        df_global["Bottom-Up Phase"] = df_global["Bottom-Up Phase"].map({1: "Yes", 0: "No"})
        df_global.to_excel(writer, index=False, sheet_name="Global Projection")
        _style_sheet(writer.sheets["Global Projection"])

        # ── Sheet 3: Regional Breakdown ───────────────────────────────────────
        regional_rows = []
        for region, df_r in proj_by_region.items():
            if df_r.empty:
                continue
            for _, row in df_r.iterrows():
                regional_rows.append({
                    "Region":              region,
                    "Year":                int(row["year"]),
                    "Capacity (GW)":       round(float(row["capacity_operating_gw"]), 2),
                    "Retirements (GW/yr)": round(float(row["retirements_this_year_gw"]), 3),
                    "Additions (GW/yr)":   round(float(row["additions_this_year_gw"]), 3),
                    "Bottom-Up Phase":     "Yes" if row["is_bottom_up"] else "No",
                })
        df_regional = pd.DataFrame(regional_rows)
        if not df_regional.empty:
            df_regional.to_excel(writer, index=False, sheet_name="Regional Breakdown")
            _style_sheet(writer.sheets["Regional Breakdown"])

        # ── Sheet 4: About ────────────────────────────────────────────────────
        about_rows = [
            ("Nuclear Capacity Dashboard — Scenario Export", ""),
            ("", ""),
            ("Generated", str(_date.today())),
            ("Scenario", active_key.title()),
            ("", ""),
            ("Model overview", ""),
            ("Phase 1 (2024–2040)", "Bottom-up unit-level simulation. Each reactor tracked individually."),
            ("Phase 2 (2040–2050)", "Top-down growth bridge. New build driven by GW/yr rate; unit retirements continue."),
            ("", ""),
            ("Data sources", ""),
            ("Reactor data",      "IAEA PRIS (© IAEA) — operating fleet, commissioning & retirement dates"),
            ("Pipeline data",     "WNA Reactor Database (© WNA) — under construction, planned, proposed units"),
            ("Benchmarks",        "IAEA RDS-1 2025 (Low/High) · IEA WEO 2024 (STEPS/APS/Low Nuclear)"),
            ("Historical capacity","IAEA RDS-2 (2005–2023)"),
            ("", ""),
            ("Disclaimer", ""),
            ("", "These projections are scenario outputs, not forecasts. They reflect the"),
            ("", "assumptions set by the user and should not be used as the basis for"),
            ("", "investment, policy, or engineering decisions without independent verification."),
        ]
        df_about = pd.DataFrame(about_rows, columns=["Field", "Detail"])
        df_about.to_excel(writer, index=False, sheet_name="About")
        ws_a = writer.sheets["About"]
        _style_sheet(ws_a, header_color="2e7d32")
        for row in ws_a.iter_rows(min_row=2):
            val = str(row[0].value or "")
            if val and not val.startswith(" ") and row[1].value == "":
                row[0].font = Font(bold=True)

    return buf.getvalue()
