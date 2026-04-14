"""
app.py — Streamlit entry point for the Nuclear Capacity Dashboard.

Run with:
    cd nuclear-dashboard
    streamlit run dashboard/app.py
"""
import streamlit as st
import pandas as pd
import io
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from config import REGIONS, DB_PATH, BASELINE_YEAR, PROJECTION_END_YEAR
from model.api import (
    get_projection, get_historical, get_benchmarks,
    get_reactor_list, get_data_vintage_string,
    write_and_run_custom_scenario, get_country_capacity,
    get_full_reactor_download, cleanup_stale_custom_scenarios,
    get_tech_projection, TECH_GROUPS,
)
from dashboard.charts import (
    chart1_global_projection,
    chart4_regional,
    chart4_technology,
    chart5_retirements,
    chart6_additions,
    chart7_country_bar,
    chart_map,
)
from dashboard.levers import render_lever_panel, ScenarioState, PRESET_DEFAULTS

# ── Page config ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Nuclear Capacity Dashboard",
    page_icon="⚛️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS: widen sidebar to ~35% ─────────────────────────────────────
st.markdown("""
<style>
    [data-testid="stSidebar"] {
        min-width: 380px;
        max-width: 440px;
    }
    [data-testid="stSidebar"] > div:first-child {
        padding-top: 1rem;
    }
    .stTabs [data-baseweb="tab-list"] { gap: 6px; }
    .stTabs [data-baseweb="tab"] { padding: 6px 14px; font-size: 13px; }
    footer { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ── Session initialisation ────────────────────────────────────────────────
# Each browser session gets its own unique custom-scenario ID so concurrent
# users never overwrite each other's DB rows.
if "_custom_scenario_id" not in st.session_state:
    import uuid
    st.session_state["_custom_scenario_id"] = f"custom_{uuid.uuid4().hex[:8]}"
    # Run cleanup once per new session (background hygiene — delete stale rows)
    cleanup_stale_custom_scenarios(max_age_hours=2)

_SESSION_CUSTOM_ID: str = st.session_state["_custom_scenario_id"]


# ── Data loading (cached) ──────────────────────────────────────────────────

@st.cache_data(ttl=300)
def load_all_projections() -> dict[str, dict[str, pd.DataFrame]]:
    scenario_ids = ["decline", "conservative", "base", "optimistic"]
    result = {}
    for sc_id in scenario_ids:
        result[sc_id] = {}
        for region in REGIONS + ["Global"]:
            result[sc_id][region] = get_projection(scenario_id=sc_id, region=region)
    return result


@st.cache_data(ttl=300)
def load_historical_all() -> dict[str, pd.DataFrame]:
    result = {}
    for region in REGIONS + ["Global"]:
        result[region] = get_historical(region=region)
    return result


@st.cache_data(ttl=300)
def load_benchmarks() -> pd.DataFrame:
    return get_benchmarks(region="Global")


@st.cache_data(ttl=300)
def load_reactors() -> pd.DataFrame:
    return get_reactor_list()


@st.cache_data(ttl=600)
def load_vintage() -> str:
    return get_data_vintage_string()


@st.cache_data(ttl=600)
def load_full_reactor_download() -> pd.DataFrame:
    return get_full_reactor_download()


@st.cache_data(ttl=300, show_spinner=False)
def _to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Data") -> bytes:
    """Convert a DataFrame to Excel bytes suitable for st.download_button.
    Cached so repeated re-runs (slider moves, tab switches) don't regenerate the file."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        from openpyxl.styles import Font, PatternFill, Alignment
        hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        ws.freeze_panes = "A2"
    return buf.getvalue()


@st.cache_data(ttl=300, show_spinner=False)
def load_tech_projection(
    scenario_id: str,
    regions_key: str,   # stringified tuple of regions — hashable cache key
    smr_post2040_share: float,
    smr_accel_start_year: int = 0,
    smr_accel_gw_per_year: float = 0.0,
) -> pd.DataFrame:
    """Cached wrapper for get_tech_projection."""
    regions = list(eval(regions_key)) if regions_key != "all" else None
    return get_tech_projection(
        scenario_id=scenario_id,
        regions=regions,
        smr_post2040_share=smr_post2040_share,
        smr_accel_start_year=smr_accel_start_year,
        smr_accel_gw_per_year=smr_accel_gw_per_year,
    )


@st.cache_data(ttl=300, show_spinner=False)
def load_country_capacity(year: int, scenario_id: str) -> pd.DataFrame:
    """Cached wrapper for get_country_capacity — avoids rebuilding retirement
    schedules and pipeline on every slider drag."""
    return get_country_capacity(year=year, scenario_id=scenario_id)


# ── Scenario export ───────────────────────────────────────────────────────

def _build_scenario_export(
    state: "ScenarioState",
    active_key: str,
    proj_global: pd.DataFrame,
    proj_by_region: dict,
) -> bytes:
    """
    Build a multi-sheet Excel export for the current scenario:
      Sheet 1 — Scenario Settings  (all lever values, fully reproducible)
      Sheet 2 — Global Projection  (year-by-year global totals)
      Sheet 3 — Regional Breakdown (year-by-year per region)
      Sheet 4 — About              (model description + disclaimer)
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import date as _date

    buf = io.BytesIO()

    # ── Helper: style a worksheet header row ──────────────────────────────
    def _style_sheet(ws, header_color="1F4E79"):
        hdr_fill = PatternFill(start_color=header_color, end_color=header_color,
                               fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
        ws.freeze_panes = "A2"

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── Sheet 1: Scenario Settings ────────────────────────────────────
        ext_labels = {
            "None":           "None — retire at baseline date",
            "HistoricalRate": "Historical Rate (~50% of fleet extended)",
            "Moderate":       "Moderate (~75% of fleet extended)",
            "MaximumAllowed": "Maximum Allowed (100%, max regulatory life)",
        }
        pipeline_desc = {
            "high":   "High — UC + Planned + Proposed all complete",
            "medium": "Medium — UC + Planned complete; Proposed excluded",
            "low":    "Low — Under Construction only",
        }
        settings_rows = [
            ("Scenario",                  active_key.title()),
            ("Export date",               str(_date.today())),
            ("",                          ""),
            ("── Life Extensions",        ""),
            ("Extension policy",          ext_labels.get(state.extension_policy_global,
                                                          state.extension_policy_global)),
            ("",                          ""),
            ("── Pipeline Realization",   ""),
            ("Large reactor preset",      pipeline_desc.get(state.large_pipeline_preset,
                                                             state.large_pipeline_preset)),
            ("  Under Construction rate", f"{state.pipeline_uc_rate*100:.0f}%"),
            ("  Planned rate",            f"{state.pipeline_planned_rate*100:.0f}%"),
            ("  Proposed rate",           f"{state.pipeline_proposed_rate*100:.0f}%"),
            ("SMR pipeline preset",       pipeline_desc.get(state.smr_pipeline_preset,
                                                             state.smr_pipeline_preset)),
            ("  SMR Under Construction",  f"{state.smr_uc_rate*100:.0f}%"),
            ("  SMR Planned",             f"{state.smr_planned_rate*100:.0f}%"),
            ("  SMR Proposed",            f"{state.smr_proposed_rate*100:.0f}%"),
            ("Construction delay adder",  f"+{state.construction_delay_adder:.0f} years"),
            ("",                          ""),
            ("── SMR Deployment",         ""),
            ("Pre-2040 SMR acceleration", f"{state.smr_accel_gw_per_year:.1f} GW/yr"
                                          + (f" from {state.smr_accel_start_year}"
                                             if state.smr_accel_gw_per_year > 0 else " (none)")),
            ("Post-2040 SMR share",       f"{state.smr_post2040_share*100:.0f}%"),
            ("",                          ""),
            ("── Post-2040 New Build",    ""),
            ("Global new build rate",     f"{state.post2040_global_growth_gw:.1f} GW/yr"),
            ("China share",               f"{state.china_post2040_gw:.1f} GW/yr"),
            ("Rest-of-world share",
             f"{max(0.0, state.post2040_global_growth_gw - state.china_post2040_gw):.1f} GW/yr"),
        ]
        df_settings = pd.DataFrame(settings_rows, columns=["Setting", "Value"])
        df_settings.to_excel(writer, index=False, sheet_name="Scenario Settings")
        ws_s = writer.sheets["Scenario Settings"]
        _style_sheet(ws_s, header_color="1F4E79")
        # Bold section headers (rows where Value is empty and Setting starts with ──)
        for row in ws_s.iter_rows(min_row=2):
            if str(row[0].value or "").startswith("──"):
                for cell in row:
                    cell.font = Font(bold=True, color="1F4E79")

        # ── Sheet 2: Global Projection ────────────────────────────────────
        df_global = proj_global[[
            "year", "capacity_operating_gw",
            "retirements_this_year_gw", "additions_this_year_gw",
            "capacity_retired_ytd_gw", "capacity_added_ytd_gw", "is_bottom_up",
        ]].copy()
        df_global.columns = [
            "Year", "Capacity (GW)", "Retirements (GW/yr)", "Additions (GW/yr)",
            "Cumul. Retired (GW)", "Cumul. Added (GW)", "Bottom-Up Phase",
        ]
        df_global["Capacity (GW)"] = df_global["Capacity (GW)"].round(2)
        df_global["Retirements (GW/yr)"] = df_global["Retirements (GW/yr)"].round(3)
        df_global["Additions (GW/yr)"] = df_global["Additions (GW/yr)"].round(3)
        df_global["Cumul. Retired (GW)"] = df_global["Cumul. Retired (GW)"].round(2)
        df_global["Cumul. Added (GW)"] = df_global["Cumul. Added (GW)"].round(2)
        df_global["Bottom-Up Phase"] = df_global["Bottom-Up Phase"].map({1: "Yes", 0: "No"})
        df_global.to_excel(writer, index=False, sheet_name="Global Projection")
        _style_sheet(writer.sheets["Global Projection"])

        # ── Sheet 3: Regional Breakdown ───────────────────────────────────
        regional_rows = []
        for region, df_r in proj_by_region.items():
            if df_r.empty:
                continue
            for _, row in df_r.iterrows():
                regional_rows.append({
                    "Region":               region,
                    "Year":                 int(row["year"]),
                    "Capacity (GW)":        round(float(row["capacity_operating_gw"]), 2),
                    "Retirements (GW/yr)":  round(float(row["retirements_this_year_gw"]), 3),
                    "Additions (GW/yr)":    round(float(row["additions_this_year_gw"]), 3),
                    "Bottom-Up Phase":      "Yes" if row["is_bottom_up"] else "No",
                })
        df_regional = pd.DataFrame(regional_rows)
        if not df_regional.empty:
            df_regional.to_excel(writer, index=False, sheet_name="Regional Breakdown")
            _style_sheet(writer.sheets["Regional Breakdown"])

        # ── Sheet 4: About ────────────────────────────────────────────────
        about_rows = [
            ("Nuclear Capacity Dashboard — Scenario Export", ""),
            ("", ""),
            ("Generated", str(_date.today())),
            ("Scenario", active_key.title()),
            ("", ""),
            ("Model overview", ""),
            ("Phase 1 (2024–2040)", "Bottom-up unit-level simulation. Each reactor tracked individually."),
            ("Phase 2 (2040–2050)", "Top-down growth bridge. New build driven by GW/yr rate; unit retirements continue."),
            ("", ""),
            ("Data sources", ""),
            ("Reactor data", "IAEA PRIS (© IAEA) — operating fleet, commissioning & retirement dates"),
            ("Pipeline data", "WNA Reactor Database (© WNA) — under construction, planned, proposed units"),
            ("Benchmarks", "IAEA RDS-1 2025 (Low/High) · IEA WEO 2024 (STEPS/APS/Low Nuclear)"),
            ("Historical capacity", "IAEA RDS-2 (2005–2023)"),
            ("", ""),
            ("Disclaimer", ""),
            ("", "These projections are scenario outputs, not forecasts. They reflect the"),
            ("", "assumptions set by the user and should not be used as the basis for"),
            ("", "investment, policy, or engineering decisions without independent verification."),
        ]
        df_about = pd.DataFrame(about_rows, columns=["Field", "Detail"])
        df_about.to_excel(writer, index=False, sheet_name="About")
        ws_a = writer.sheets["About"]
        _style_sheet(ws_a, header_color="2e7d32")
        # Bold the section labels
        for row in ws_a.iter_rows(min_row=2):
            val = str(row[0].value or "")
            if val and not val.startswith(" ") and row[1].value == "":
                row[0].font = Font(bold=True)

    return buf.getvalue()


# ── Geography-aware aggregation helpers ───────────────────────────────────

def _sum_projections(proj_dict: dict, regions: list) -> pd.DataFrame:
    """Sum per-region projection DataFrames across the given regions.
    Uses year-based groupby so misaligned or differently-sized DataFrames
    cannot cause silent positional errors."""
    sum_cols = ["capacity_operating_gw", "retirements_this_year_gw", "additions_this_year_gw"]
    dfs = [proj_dict[r][["year", "is_bottom_up"] + sum_cols]
           for r in regions if r in proj_dict and not proj_dict[r].empty]
    if not dfs:
        return pd.DataFrame()
    combined = pd.concat(dfs)
    # is_bottom_up is identical across regions for any given year — take max (1 > 0)
    ib = combined.groupby("year")["is_bottom_up"].max().reset_index()
    summed = combined.groupby("year")[sum_cols].sum().reset_index()
    return summed.merge(ib, on="year")


def _sum_historical(hist_dict: dict, regions: list) -> pd.DataFrame:
    """Sum per-region historical DataFrames across the given regions."""
    dfs = [hist_dict[r][["year", "capacity_gw"]]
           for r in regions if r in hist_dict and not hist_dict[r].empty]
    if not dfs:
        return pd.DataFrame()
    return pd.concat(dfs).groupby("year")["capacity_gw"].sum().reset_index()


# ── Sidebar lever panel ────────────────────────────────────────────────────
state, lever_updated = render_lever_panel()

# ── Load data ──────────────────────────────────────────────────────────────
all_projections = load_all_projections()
historical_all = load_historical_all()
benchmarks = load_benchmarks()
reactors = load_reactors()
vintage = load_vintage()

sc_id = state.scenario_id

# ── Custom lever recomputation ─────────────────────────────────────────────
# When Update is clicked, check if lever values differ from the selected preset.
# If so, recompute a "custom" scenario and overlay it on the active chart.
# Derive comparison defaults from the canonical PRESET_DEFAULTS in levers.py.
# smr_post2040_share is stored as integer percent in levers.py; convert to float fraction here.
_preset_defaults = {
    k: {**v, "smr_post2040_share": v["smr_post2040_share"] / 100.0}
    for k, v in PRESET_DEFAULTS.items()
}

def _levers_match_preset(state: ScenarioState, preset_id: str) -> bool:
    pd_ = _preset_defaults.get(preset_id, {})
    return (
        state.extension_policy_global == pd_.get("extension_policy_global") and
        state.large_pipeline_preset == pd_.get("large_pipeline_preset") and
        state.smr_pipeline_preset == pd_.get("smr_pipeline_preset") and
        abs(state.construction_delay_adder - pd_.get("construction_delay_adder", 0)) < 0.01 and
        # acceleration: only check rate (if rate=0, start year is irrelevant)
        abs(state.smr_accel_gw_per_year - pd_.get("smr_accel_gw_per_year", 0)) < 0.01 and
        abs(state.smr_post2040_share - pd_.get("smr_post2040_share", 0.20)) < 0.01 and
        abs(state.post2040_global_growth_gw - pd_.get("post2040_global_growth_gw", 0)) < 0.1 and
        abs(state.china_post2040_gw - pd_.get("china_post2040_gw", 0)) < 0.1
    )

custom_projection: dict | None = st.session_state.get("_custom_projection")

# Clear custom if user switched presets (even without clicking Update)
if st.session_state.get("_custom_for_scenario") != sc_id:
    st.session_state.pop("_custom_projection", None)
    custom_projection = None

if lever_updated:
    if _levers_match_preset(state, sc_id):
        # Levers match preset — use precomputed data, clear any custom
        st.session_state.pop("_custom_projection", None)
        custom_projection = None
    else:
        # Custom settings — recompute using session-unique scenario ID
        with st.spinner("Computing custom projection…"):
            custom_projection = write_and_run_custom_scenario(
                extension_policy=state.extension_policy_global,
                pipeline_uc_rate=state.pipeline_uc_rate,
                pipeline_planned_rate=state.pipeline_planned_rate,
                pipeline_proposed_rate=state.pipeline_proposed_rate,
                construction_delay_adder=state.construction_delay_adder,
                post2040_global_gw=state.post2040_global_growth_gw,
                smr_uc_rate=state.smr_uc_rate,
                smr_planned_rate=state.smr_planned_rate,
                smr_proposed_rate=state.smr_proposed_rate,
                china_post2040_gw=state.china_post2040_gw,
                smr_accel_start_year=state.smr_accel_start_year,
                smr_accel_gw_per_year=state.smr_accel_gw_per_year,
                scenario_id=_SESSION_CUSTOM_ID,
            )
        st.session_state["_custom_projection"] = custom_projection
        st.session_state["_custom_for_scenario"] = sc_id

# Use custom projection if available, else precomputed
if custom_projection is not None:
    proj_global = custom_projection["Global"]
    _proj_source = "Custom"
    # DB scenario ID for queries that go back to the database (e.g. country capacity)
    _active_db_id = _SESSION_CUSTOM_ID
else:
    proj_global = all_projections[sc_id]["Global"]
    _proj_source = sc_id
    _active_db_id = sc_id

hist_global = historical_all["Global"]

# Geography filter: use state.selected_regions
sel_regions = [r for r in REGIONS if r in state.selected_regions]
if not sel_regions:
    sel_regions = list(REGIONS)
geo_filtered = len(sel_regions) < len(REGIONS)

# Per-region projections and historical (already filtered to sel_regions)
_proj_source_dict = custom_projection if custom_projection is not None else all_projections[sc_id]
proj_by_region = {r: _proj_source_dict[r] for r in sel_regions if r in _proj_source_dict}
hist_by_region = {r: historical_all[r] for r in sel_regions if r in historical_all}

# Reactors filtered to selected regions (for pipeline funnel + fleet metrics)
filtered_reactors = reactors[reactors["region"].isin(sel_regions)] if geo_filtered else reactors

# Geography-aware global view:
# When filter is active, sum the selected-region projections/historical;
# otherwise use the pre-aggregated Global row (faster, exact).
if geo_filtered:
    proj_global_display = _sum_projections(_proj_source_dict, sel_regions)
    hist_global_display = _sum_historical(historical_all, sel_regions)
else:
    proj_global_display = proj_global
    hist_global_display = hist_global

# Build comparison projections — each also filtered to sel_regions when active
# active_key is used for chart labels/colors — always "custom" for custom projections
active_key = "custom" if custom_projection is not None else sc_id
if geo_filtered:
    compare_projs = {
        cid: _sum_projections(all_projections[cid], sel_regions)
        for cid in state.compare_scenarios
        if cid in all_projections
    }
else:
    compare_projs = {
        cid: all_projections[cid]["Global"]
        for cid in state.compare_scenarios
        if cid in all_projections
    }
compare_projs[active_key] = proj_global_display

# ── Sidebar: scenario export button ───────────────────────────────────────
st.sidebar.markdown("---")
st.sidebar.markdown("### 💾 Export Scenario")
_export_filename = f"nuclear_scenario_{active_key}_{pd.Timestamp.today().strftime('%Y-%m-%d')}.xlsx"
st.sidebar.download_button(
    label="⬇ Download Scenario + Settings",
    data=_build_scenario_export(state, active_key, proj_global_display, proj_by_region),
    file_name=_export_filename,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
    help=(
        "Download the current scenario as a 4-sheet Excel file: "
        "lever settings, global projection, regional breakdown, and model description."
    ),
)
if custom_projection is not None:
    st.sidebar.caption("📌 Exporting custom lever settings + projections.")
else:
    st.sidebar.caption(f"📌 Exporting **{active_key.title()}** preset + projections.")


# ── Header ─────────────────────────────────────────────────────────────────
st.markdown("## ⚛️ Global Nuclear Capacity Dashboard")
col_h1, col_h2, col_h3, col_h4 = st.columns(4)

def _gw(df, default=0.0):
    return float(df["capacity_operating_gw"].iloc[0]) if not df.empty else default

kpi_suffix = " ★" if geo_filtered else ""   # star signals filtered view
kpi_help   = " (selected regions only)" if geo_filtered else ""

baseline_row = proj_global_display[proj_global_display["year"] == BASELINE_YEAR]
row_2030     = proj_global_display[proj_global_display["year"] == 2030]
row_2040     = proj_global_display[proj_global_display["year"] == 2040]
end_row      = proj_global_display[proj_global_display["year"] == PROJECTION_END_YEAR]

with col_h1:
    st.metric(f"2024 Baseline{kpi_suffix}", f"{_gw(baseline_row):.0f} GW",
              help=f"Current operating capacity (PRIS 2024){kpi_help}")
with col_h2:
    val30 = _gw(row_2030)
    st.metric(f"2030 Projection{kpi_suffix}", f"{val30:.0f} GW",
              delta=f"{val30-_gw(baseline_row):+.0f} GW vs 2024",
              help=kpi_help.strip() or None)
with col_h3:
    val40 = _gw(row_2040)
    st.metric(f"2040 (Transition){kpi_suffix}", f"{val40:.0f} GW",
              delta=f"{val40-_gw(baseline_row):+.0f} GW vs 2024",
              help=kpi_help.strip() or None)
with col_h4:
    val50 = _gw(end_row)
    st.metric(f"2050 Projection{kpi_suffix}", f"{val50:.0f} GW",
              delta=f"{val50-_gw(baseline_row):+.0f} GW vs 2024",
              help=kpi_help.strip() or None)

_dl_col1, _dl_col2 = st.columns([6, 1])
with _dl_col2:
    _reactor_dl_df = load_full_reactor_download()
    st.download_button(
        label="⬇ Full Reactor List",
        data=_to_excel_bytes(_reactor_dl_df, "All Reactors"),
        file_name="nuclear_reactor_list.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help=f"Download all {len(_reactor_dl_df):,} reactors (operating, pipeline, long-term shutdown)",
        use_container_width=True,
    )

st.markdown("---")

# Geography filter banner
if geo_filtered:
    excluded = [r for r in REGIONS if r not in sel_regions]
    st.info(
        f"🌍 **Geography filter active** — showing {len(sel_regions)} of {len(REGIONS)} regions. "
        f"All charts, KPIs ★, and historical data reflect selected regions only. "
        f"Excluded: {', '.join(excluded)}."
    )

# ── Chart tabs ─────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "📈 Global Projection",
    "📊 Capacity Breakdown",
    "📉 Retirements",
    "➕ New Additions",
    "🏳️ Country Snapshot",
    "🗺️ World Map",
])


# ── Tab 1: Global Projection ───────────────────────────────────────────────
with tab1:
    # Hide benchmarks when a geography filter is active — global benchmarks
    # are not comparable to a regional subset of capacity.
    if geo_filtered:
        show_benchmarks = {k: False for k in ["IAEA Low", "IAEA High", "IEA STEPS", "IEA APS", "IEA Low Nuclear"]}
    else:
        show_benchmarks = {
            "IAEA Low":        state.show_iaea_low,
            "IAEA High":       state.show_iaea_high,
            "IEA STEPS":       state.show_iea_steps,
            "IEA APS":         state.show_iea_aps,
            "IEA Low Nuclear": state.show_iea_low_nuclear,
        }
    if custom_projection is not None:
        st.info("Custom lever settings applied. Preset lines available via comparison selector.")
    if geo_filtered:
        st.caption(f"📊 Showing sum of {len(sel_regions)} selected regions — benchmark lines hidden (global comparisons not applicable to regional subsets).")

    fig1 = chart1_global_projection(
        projections=compare_projs,
        historical=hist_global_display if state.show_historical else pd.DataFrame(),
        benchmarks=benchmarks,
        active_scenario=active_key,
        compare_scenarios=state.compare_scenarios,
        show_historical=state.show_historical,
        show_benchmarks=show_benchmarks,
        show_transition_marker=state.show_transition_marker,
    )
    _t1c1, _t1c2 = st.columns([8, 1])
    with _t1c1:
        st.plotly_chart(fig1, use_container_width=True)
        st.caption(
            "**2040 transition:** projections switch from bottom-up (individual reactor data — "
            "retirements + announced pipeline only) to top-down (global new-build rate lever). "
            "**Benchmark gap:** IEA/IAEA reference scenarios assume 100–200+ GW of capacity not yet "
            "announced or under licence — this model includes only reactors with a known construction "
            "or licence decision. The gap is by design and expected to widen through 2040. "
            "IEA STEPS/APS 2030 markers (~513/551 GW) serve as near-term cross-checks."
        )
        if active_key in ("decline", "custom") and (
            active_key == "decline"
            or state.extension_policy_global == "AcceleratedRetirement"
        ):
            st.caption(
                "ℹ️ **No New Extensions policy:** Existing approved licenses are honored in full; "
                "zero new extension rounds are granted beyond what regulators have already approved. "
                "US reactors (60-yr licenses) retire through 2029–2045; Japanese reactors follow "
                "their current approved terms. The gradual capacity decline reflects this staggered "
                "retirement schedule — not an immediate phase-out."
            )
    with _t1c2:
        _t1_dl = proj_global_display[["year", "capacity_operating_gw",
                                      "retirements_this_year_gw", "additions_this_year_gw"]].copy()
        _t1_dl.columns = ["Year", "Capacity (GW)", "Retirements (GW)", "Additions (GW)"]
        _t1_dl = _t1_dl.round(2)
        st.download_button("⬇ Data", data=_to_excel_bytes(_t1_dl, "Global Projection"),
                           file_name="global_projection.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

    with st.expander("Show projection data table"):
        _tbl_src = proj_global_display if not proj_global_display.empty else proj_global
        display_df = _tbl_src[["year", "capacity_operating_gw",
                                "retirements_this_year_gw",
                                "additions_this_year_gw",
                                "is_bottom_up"]].copy()
        display_df.columns = ["Year", "Capacity (GW)", "Retirements (GW)", "Additions (GW)", "Bottom-Up"]
        display_df["Capacity (GW)"] = display_df["Capacity (GW)"].round(1)
        display_df["Retirements (GW)"] = display_df["Retirements (GW)"].round(2)
        display_df["Additions (GW)"] = display_df["Additions (GW)"].round(2)
        display_df["Bottom-Up"] = display_df["Bottom-Up"].map({1: "Yes", 0: "No"})
        st.dataframe(display_df, use_container_width=True, hide_index=True)


# ── Tab 2: Capacity Breakdown ──────────────────────────────────────────────
with tab2:
    breakdown_by = st.radio(
        "View by",
        options=["regional", "technology"],
        format_func=lambda x: "Geography (by region)" if x == "regional" else "Technology (PWR / BWR / PHWR / SMR / Other)",
        horizontal=True,
        key="breakdown_split",
    )

    _t4c1, _t4c2 = st.columns([8, 1])

    if breakdown_by == "regional":
        fig4 = chart4_regional(
            projections_by_region=proj_by_region,
            historical_by_region=hist_by_region,
            regions=sel_regions,
            scenario_id=active_key,
            show_historical=state.show_historical,
        )
        with _t4c1:
            st.plotly_chart(fig4, use_container_width=True)
            st.caption(
                "Historical regional totals (shaded area, pre-2025) are derived from 5-year IAEA PRIS "
                "country snapshots with linear interpolation between survey years. The Global total "
                "(Tab 1) uses annual IAEA statistics and may differ from the regional sum in years "
                "affected by sudden capacity changes (e.g., Japan post-Fukushima 2011–2014) or "
                "countries not reported individually by IAEA (e.g., Taiwan). "
                "Post-2025 projections are scenario-specific."
            )
        with _t4c2:
            _t4_rows = []
            for _r in sel_regions:
                _df_r = _proj_source_dict.get(_r, pd.DataFrame())
                if _df_r.empty:
                    continue
                for _, _row in _df_r.iterrows():
                    _t4_rows.append({
                        "Region": _r,
                        "Year": int(_row["year"]),
                        "Capacity (GW)": round(float(_row["capacity_operating_gw"]), 2),
                        "Retirements (GW)": round(float(_row["retirements_this_year_gw"]), 2),
                        "Additions (GW)": round(float(_row["additions_this_year_gw"]), 2),
                    })
            _t4_dl = pd.DataFrame(_t4_rows)
            st.download_button("⬇ Data", data=_to_excel_bytes(_t4_dl, "Regional Breakdown"),
                               file_name="regional_breakdown.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

        with st.expander("Show regional capacity table (2024 / 2030 / 2040 / 2050)"):
            def _val(df_r, yr):
                row = df_r[df_r["year"] == yr]
                return float(row["capacity_operating_gw"].iloc[0]) if not row.empty else 0.0

            rows = []
            for region in REGIONS:
                df_r = _proj_source_dict.get(region, pd.DataFrame())
                if df_r.empty:
                    continue
                rows.append({
                    "Region": region,
                    "2024 (GW)": round(_val(df_r, 2024), 1),
                    "2030 (GW)": round(_val(df_r, 2030), 1),
                    "2040 (GW)": round(_val(df_r, 2040), 1),
                    "2050 (GW)": round(_val(df_r, 2050), 1),
                })
            if rows:
                tbl = pd.DataFrame(rows)
                tbl.loc[len(tbl)] = {
                    "Region": "GLOBAL",
                    "2024 (GW)": round(_gw(proj_global[proj_global["year"] == 2024]), 1),
                    "2030 (GW)": round(_gw(proj_global[proj_global["year"] == 2030]), 1),
                    "2040 (GW)": round(_gw(proj_global[proj_global["year"] == 2040]), 1),
                    "2050 (GW)": round(_gw(proj_global[proj_global["year"] == 2050]), 1),
                }
                st.dataframe(tbl, use_container_width=True, hide_index=True)

    else:
        # Technology breakdown
        _regions_key = str(tuple(sorted(sel_regions))) if geo_filtered else "all"
        with st.spinner("Computing technology breakdown…"):
            tech_df = load_tech_projection(
                scenario_id=_active_db_id,
                regions_key=_regions_key,
                smr_post2040_share=state.smr_post2040_share,
                smr_accel_start_year=state.smr_accel_start_year if state.smr_accel_gw_per_year > 0 else 0,
                smr_accel_gw_per_year=state.smr_accel_gw_per_year,
            )
        st.caption(
            "Post-2040 technology split is estimated: SMR share from the lever; "
            "PWR/BWR/PHWR/Other proportions extrapolated from the 2040 fleet composition."
        )
        fig4t = chart4_technology(
            tech_df=tech_df,
            historical=hist_global_display if state.show_historical else pd.DataFrame(),
            scenario_id=active_key,
            show_historical=state.show_historical,
        )
        with _t4c1:
            st.plotly_chart(fig4t, use_container_width=True)
        with _t4c2:
            _t4t_dl = tech_df.pivot_table(
                index="year", columns="tech_group", values="capacity_gw"
            ).reset_index()
            _t4t_dl.columns.name = None
            _t4t_dl = _t4t_dl.round(2)
            st.download_button("⬇ Data",
                               data=_to_excel_bytes(_t4t_dl, "Technology Breakdown"),
                               file_name="technology_breakdown.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

        with st.expander("Show technology capacity table (2024 / 2030 / 2040 / 2050)"):
            _tech_pivot = tech_df[tech_df["year"].isin([2024, 2030, 2040, 2050])].pivot_table(
                index="tech_group", columns="year", values="capacity_gw"
            ).reset_index()
            _tech_pivot.columns = ["Technology"] + [str(c) + " (GW)" for c in _tech_pivot.columns[1:]]
            _tech_pivot = _tech_pivot.round(1)
            st.dataframe(_tech_pivot, use_container_width=True, hide_index=True)


# ── Tab 3: Retirements ─────────────────────────────────────────────────────
with tab3:
    st.caption(
        "Annual retirements shown as negative values. Stacked by region — "
        "areas reflect gross capacity removed each year."
    )
    fig5 = chart5_retirements(
        projections_by_region=proj_by_region,
        regions=sel_regions,
        scenario_id=active_key,
    )
    _t5c1, _t5c2 = st.columns([8, 1])
    with _t5c1:
        st.plotly_chart(fig5, use_container_width=True)
    with _t5c2:
        _t5_rows = []
        for _r in sel_regions:
            _df_r = _proj_source_dict.get(_r, pd.DataFrame())
            if _df_r.empty:
                continue
            for _, _row in _df_r.iterrows():
                if _row["year"] <= 2024:
                    continue
                _t5_rows.append({
                    "Region": _r,
                    "Year": int(_row["year"]),
                    "Retirements (GW)": round(float(_row["retirements_this_year_gw"]), 3),
                })
        _t5_dl = pd.DataFrame(_t5_rows)
        st.download_button("⬇ Data", data=_to_excel_bytes(_t5_dl, "Retirements"),
                           file_name="retirements_by_region.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)


# ── Tab 4: New Additions ───────────────────────────────────────────────────
with tab4:
    split_by = st.radio(
        "View by",
        options=["geography", "technology"],
        format_func=lambda x: "Geography (by region)" if x == "geography" else "Technology (SMR vs Large)",
        horizontal=True,
        key="additions_split",
    )
    # For technology split: compute tech_df so SMR fraction uses scenario-adjusted
    # pipeline (delay adder + realization rates) rather than raw reactor data.
    _tech_df_for_chart6: pd.DataFrame | None = None
    if split_by == "technology":
        _regions_key_t3 = str(tuple(sorted(sel_regions))) if geo_filtered else "all"
        with st.spinner("Computing technology breakdown…"):
            _tech_df_for_chart6 = load_tech_projection(
                scenario_id=_active_db_id,
                regions_key=_regions_key_t3,
                smr_post2040_share=state.smr_post2040_share,
                smr_accel_start_year=state.smr_accel_start_year if state.smr_accel_gw_per_year > 0 else 0,
                smr_accel_gw_per_year=state.smr_accel_gw_per_year,
            )
    # Warning: planned reactors with no announced construction date are excluded
    _no_date = reactors[
        (reactors["status"] == "Planned") & (reactors["expected_online_year"].isna())
    ]
    if not _no_date.empty:
        _nd_gw = _no_date["net_capacity_mw"].sum() / 1000
        _nd_countries = _no_date["country"].value_counts().head(4)
        _nd_summary = ", ".join(f"{c} ({n})" for c, n in _nd_countries.items())
        st.info(
            f"ℹ️ **{len(_no_date)} Planned reactors ({_nd_gw:.1f} GW) have no announced "
            f"construction-start date and are excluded from all projections.** "
            f"Largest groups: {_nd_summary}. "
            f"These units will appear in projections once an expected online year is confirmed."
        )
    st.caption(
        "Annual new capacity additions. Post-2040 SMR share set by the SMR lever. "
        "Pre-2040 SMR share derived from the scenario-adjusted pipeline "
        "(technology split) or summed by region (geography split)."
    )
    fig6 = chart6_additions(
        projections_by_region=proj_by_region,
        reactors=reactors,
        regions=sel_regions,
        scenario_id=active_key,
        split_by=split_by,
        smr_post2040_share=state.smr_post2040_share,
        tech_df=_tech_df_for_chart6,
    )
    _t6c1, _t6c2 = st.columns([8, 1])
    with _t6c1:
        st.plotly_chart(fig6, use_container_width=True)
    with _t6c2:
        _t6_rows = []
        for _r in sel_regions:
            _df_r = _proj_source_dict.get(_r, pd.DataFrame())
            if _df_r.empty:
                continue
            for _, _row in _df_r.iterrows():
                if _row["year"] <= 2024:
                    continue
                _t6_rows.append({
                    "Region": _r,
                    "Year": int(_row["year"]),
                    "Additions (GW)": round(float(_row["additions_this_year_gw"]), 3),
                })
        _t6_dl = pd.DataFrame(_t6_rows)
        st.download_button("⬇ Additions", data=_to_excel_bytes(_t6_dl, "New Additions"),
                           file_name="new_additions_by_region.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
        # Pipeline reactor list download
        _pipe_dl = filtered_reactors[filtered_reactors["status"].isin(
            ["UnderConstruction", "Planned", "Proposed"])].copy()
        _pipe_cols = ["name", "country", "region", "status", "net_capacity_mw",
                      "expected_online_year", "reactor_type", "is_smr"]
        _pipe_cols = [c for c in _pipe_cols if c in _pipe_dl.columns]
        _pipe_dl = _pipe_dl[_pipe_cols].sort_values(["status", "expected_online_year", "country"])
        _pipe_dl.columns = ["Name", "Country", "Region", "Status", "Capacity (MW)",
                            "Expected Online Year", "Reactor Type", "SMR"]
        st.download_button("⬇ Pipeline", data=_to_excel_bytes(_pipe_dl, "Pipeline Reactors"),
                           file_name="pipeline_reactors.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)


# ── Tab 5: Country Snapshot ────────────────────────────────────────────────
with tab5:
    col71, col72 = st.columns([1, 3])
    with col71:
        snapshot_year = st.select_slider(
            "Snapshot year",
            options=list(range(2005, 2051)),
            value=2026,
            key="country_snapshot_year",
        )
    with col72:
        st.caption(
            f"Country-level nuclear capacity for **{snapshot_year}** under the "
            f"**{active_key.title()}** scenario. Operating fleet + pipeline additions "
            f"expected online by that year (weighted by realization rates)."
        )

    with st.spinner("Loading country data…"):
        country_df = load_country_capacity(year=snapshot_year, scenario_id=_active_db_id)

    fig7 = chart7_country_bar(
        country_df=country_df,
        year=snapshot_year,
        scenario_id=active_key,
        regions=sel_regions,
    )
    _t7c1, _t7c2 = st.columns([8, 1])
    with _t7c1:
        st.plotly_chart(fig7, use_container_width=True)
    with _t7c2:
        _t7_dl = country_df.copy()
        if sel_regions:
            _t7_dl = _t7_dl[_t7_dl["region"].isin(sel_regions)]
        _t7_dl = _t7_dl[_t7_dl["total_gw"] > 0].sort_values("total_gw", ascending=False)
        _t7_dl.columns = [c.replace("_", " ").title() for c in _t7_dl.columns]
        st.download_button("⬇ Data", data=_to_excel_bytes(_t7_dl, f"Country Snapshot {snapshot_year}"),
                           file_name=f"country_snapshot_{snapshot_year}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

    with st.expander("Show country data table"):
        tbl7 = country_df.copy()
        if sel_regions:
            tbl7 = tbl7[tbl7["region"].isin(sel_regions)]
        tbl7 = tbl7[tbl7["total_gw"] > 0].sort_values("total_gw", ascending=False)
        tbl7.columns = [c.replace("_", " ").title() for c in tbl7.columns]
        st.dataframe(tbl7, use_container_width=True, hide_index=True)


# ── Tab 6: World Map ──────────────────────────────────────────────────────
with tab6:
    _mc1, _mc2, _mc3 = st.columns([2, 2, 5])
    with _mc1:
        map_year = st.select_slider(
            "Year",
            options=list(range(2005, 2051)),
            value=2026,
            key="map_year",
        )
    with _mc2:
        map_metric = st.radio(
            "Show",
            options=["total_gw", "operating_gw", "pipeline_gw"],
            format_func=lambda x: {
                "total_gw": "Total (operating + pipeline)",
                "operating_gw": "Operating only",
                "pipeline_gw": "Pipeline only",
            }[x],
            key="map_metric",
        )
    with _mc3:
        if map_year < 2024:
            st.caption(
                f"Historical view ({map_year}): showing reactors operating based on "
                "actual commissioning/retirement dates. Retired reactors not in the "
                "current database are excluded."
            )
        else:
            st.caption(
                f"Projection view ({map_year}): operating fleet accounting for "
                "retirements + pipeline additions weighted by realization rates "
                f"under the **{active_key.title()}** scenario."
            )

    with st.spinner("Loading map data…"):
        map_country_df = load_country_capacity(year=map_year, scenario_id=_active_db_id)

    if geo_filtered:
        map_country_df = map_country_df[map_country_df["region"].isin(sel_regions)]

    fig_map = chart_map(
        country_df=map_country_df,
        year=map_year,
        metric=map_metric,
        scenario_id=active_key,
    )
    _tm1, _tm2 = st.columns([8, 1])
    with _tm1:
        st.plotly_chart(fig_map, use_container_width=True)
    with _tm2:
        _map_dl = map_country_df[map_country_df[map_metric] > 0].copy()
        _map_dl = _map_dl.sort_values(map_metric, ascending=False)
        _map_dl.columns = [c.replace("_", " ").title() for c in _map_dl.columns]
        st.download_button(
            "⬇ Data",
            data=_to_excel_bytes(_map_dl, f"Map {map_year}"),
            file_name=f"world_map_{map_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


# ── Footer & Disclaimer ────────────────────────────────────────────────────
st.markdown("---")

with st.expander("ℹ️ About this dashboard & methodology", expanded=False):
    st.markdown("""
**What this is**

An interactive scenario modelling tool for global nuclear power capacity from 2024 to 2050.
It is intended for research and educational purposes. Charts and projections are
**scenario outputs, not predictions** — they reflect the assumptions you set via the levers,
not a forecast of what will happen.

**Model structure**

The model operates in two phases:
- **2024–2040 (bottom-up):** Unit-level simulation. Each reactor is tracked individually —
  retirement dates are computed from country-specific licensing rules and the chosen extension
  policy; pipeline reactors are included weighted by their realization probability and any
  construction delay you apply.
- **2040–2050 (top-down):** A growth-rate bridge. Unit-level retirements continue, but new
  additions are driven by the global GW/yr rate you set, distributed across regions using
  calibrated weights.

**Key assumptions & limitations**

- Reactor retirement dates are derived from country licensing rules (Tier 2) or declared
  shutdown dates (Tier 1). Where neither is available, a 60-year default life is assumed.
- Pipeline realization rates are deterministic probability weights, not stochastic simulations.
- Post-2040 new build is a smooth top-down rate and does not model project-level uncertainty.
- Historical data (2005–2023) is sourced from IAEA RDS-2; it reflects observed capacity,
  not model outputs.
- Reactors decommissioned before 2024 are not in the database; historical country-level
  figures for earlier years may be understated as a result.

**Data sources & attribution**

| Source | Coverage | Terms |
|--------|----------|-------|
| [IAEA PRIS](https://pris.iaea.org/) | Operating & pipeline reactors, commissioning/retirement dates | © IAEA — data used with attribution; not for commercial redistribution |
| [IAEA RDS-1 2025](https://www.iaea.org/publications/15510/nuclear-power-reactors-in-the-world) | Benchmark scenarios (Low / High) | © IAEA |
| [IAEA RDS-2](https://www.iaea.org/publications/15510/nuclear-power-reactors-in-the-world) | Historical capacity 2005–2023 | © IAEA |
| [IEA World Energy Outlook 2024](https://www.iea.org/reports/world-energy-outlook-2024) | Benchmark scenarios (STEPS / APS / Low Nuclear) | © IEA — used for reference |
| [WNA Reactor Database](https://www.world-nuclear.org/) | Pipeline supplemental data | © WNA |

**Disclaimer**

This tool is provided for informational and research purposes only. The authors make no
representations as to the accuracy, completeness, or suitability of the projections for
any particular purpose. Scenario outputs should not be used as the basis for investment,
policy, or engineering decisions without independent verification.
    """)

st.caption(
    f"Data: {vintage}  ·  "
    "Model: bottom-up unit-level (2024–2040) → top-down growth bridge (2040–2050)  ·  "
    "For research & educational use only — scenario outputs, not forecasts"
)
